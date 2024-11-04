from datetime import datetime
from openpyxl import load_workbook
import textwrap
import os

static_path = f"{os.path.dirname(os.path.abspath(__file__))}"


def paste_to_order(data):
    print_border = '   ___'
    model, package, breakage, name, phone_num, add1, add2, add3 =\
        data[0], data[1], data[2], data[3], data[4], data[5], data[6] + print_border, data[7] + print_border
    wb = load_workbook(f"{static_path}/excel_files/examples/order.xlsx")
    ws = wb.worksheets[0]
    short_date = datetime.now().strftime('%Y-%m-%d %H.%M.%S')
    string_width = 25
    ws["A2"], ws["A11"], ws["A13"], ws["A15"], ws["A16"], ws["A17"], ws["A18"] =\
        model, name, phone_num, short_date, add1, add2, add3
    package = textwrap.fill(package, string_width).split("\n")
    ws["A4"] = package[0]
    if len(package) >= 2:
        ws["A5"] = package[1]
    breakage = textwrap.fill(breakage, string_width).split("\n")
    ws["A7"] = breakage[0]
    if len(breakage) >= 2:
        ws["A8"] = breakage[1]
    if len(breakage) >= 3:
        ws["A9"] = breakage[2]
    wb.save(f"{static_path}/excel_files/saved_xlsx/{name}_{short_date}.xlsx".replace(' ', '_'))
    return f"{static_path}/excel_files/saved_xlsx/{name}_{short_date}.xlsx".replace(' ', '_')


def paste_to_warranty(data):
    model, name, phone_num, slug, break_fix, price, warranty =\
        data[0], data[1], data[2], data[3], data[4], data[5], data[6]
    wb = load_workbook(f"{static_path}/excel_files/examples/warranty.xlsx")
    ws = wb.worksheets[0]
    short_date = datetime.now().strftime('%Y-%m-%d %H.%M.%S')
    string_width = 25
    price = str(price) + ' грн.'
    ws["A2"], ws["A10"], ws["A12"], ws["A15"], ws["A17"] =\
        model, name, phone_num, short_date, price
    if warranty != '0 міс.':
        ws["A8"] = warranty
    else:
        ws["A7"] = ws["A8"] = ' '
        ws["A19"] = 'ТАЛОН'
        ws["A20"] = 'ВИДАЧІ'
    break_fix = textwrap.fill(break_fix, string_width).split("\n")
    ws["A4"] = break_fix[0]
    if len(break_fix) >= 2:
        ws["A5"] = break_fix[1]
    if len(break_fix) >= 3:
        ws["A6"] = break_fix[2]
    wb.save(f"{static_path}/excel_files/saved_xlsx/warr_{name}_{short_date}.xlsx".replace(' ', '_'))
    return f"{static_path}/excel_files/saved_xlsx/warr_{name}_{short_date}.xlsx".replace(' ', '_')




